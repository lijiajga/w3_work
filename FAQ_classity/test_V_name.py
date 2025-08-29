流程 =[]
标准规范 =[]
业务操作指导书 =[]
管理手册 =[]
FAQ_message =[]
import pandas as pd
import os
import shutil
import openpyxl

FAQ_count = 0
liucheng_count = 0


def find_V_row(name):
    last_v_index = name.rfind('V')
    result = None
    # 如果找到了'V'，截取从该位置到开头的字符串
    if last_v_index != -1:
        result = name[:last_v_index]
    if last_v_index == -1:
        result = name
    if '（试用）' in result:
        result = result.replace('（试用）', '')
    if '(试用)' in result: 
        result = result.replace('(试用)', '')
    if '/' in result:
        result = result.replace('/', '')
    if '(1)' in result:
        result = result.replace('(1)', '')
    if '（1）' in result:
        result = result.replace('（1）', '')
    if '（试行）' in result:
        result = result.replace('（试行）', '')
    if '(试行)' in result:
        result = result.replace('(试行)', '')           
    return result

#是否包含'V'
def find_V(name):
    last_v_index = name.rfind('V')
    result = None
    # 如果找到了'V'，截取从该位置到开头的字符串
    if last_v_index != -1:
        result = name[:last_v_index]
    if last_v_index == -1:
        result = os.path.splitext(name)[0]
    if '（试用）' in result:
        result = result.replace('（试用）', '')
    if '(试用)' in result: 
        result = result.replace('(试用)', '')
    if '/' in result:
        result = result.replace('/', '')
    if '(1)' in result:
        result = result.replace('(1)', '')
    if '（1）' in result:
        result = result.replace('（1）', '')    
    if '（试行）' in result:
        result = result.replace('（试行）', '')
    if '(试行)' in result:
        result = result.replace('(试行)', '')
    if ' ' in result:
        result = result.replace(' ', '')            
    return result

#复制文件
def copy_file(input_path,filename,output_path):
    shutil.copy(os.path.join(input_path, filename), output_path)



#判断w3文件的类型
def get_all_list():
    file_path = r'D:\实体文件_汇总\w3文件收集规则\ISC+知识分析0716_v2.xlsx'
    sheet_name = '线上线下收集模板'
    data = pd.read_excel(file_path, sheet_name=sheet_name)
    data = data[data['归档位置'] == 'W3']
    for index, row in data.iterrows():
        if row['小类'] == '流程':
            流程.append(find_V_row(row['文档名称（当前）']))

        elif row['小类'] in ['Checklist', 'KCP', 'SOD','模板','行政发文']:
            FAQ_message.append(find_V_row(row['文档名称（当前）']))

        elif row['小类'] == '标准规范':
            标准规范.append(find_V_row(row['文档名称（当前）']))

        elif row['小类'] == '业务操作指导书':
            业务操作指导书.append(find_V_row(row['文档名称（当前）']))

        elif row['小类'] == '管理手册':
            管理手册.append(find_V_row(row['文档名称（当前）']))

#文件归档到指定的文件夹中
def file_classify():
    # 定义文件夹路径
    folder_path = r'D:\Pan_xfusion_worksapce\1.《《任务》》\7月\1.知识库上传\7.w3\w3文件_总'
    # 获取文件夹中的所有文件名
    w3_file_names = os.listdir(folder_path)

    for file in w3_file_names:
        if find_V(file) in 流程:
            copy_file(folder_path,file,r'D:\Pan_xfusion_worksapce\1.《《任务》》\7月\1.知识库上传\7.w3\w3_知识库(强匹配)\ISC+通用-流程文件-内部公开')
        
        if find_V(file) in 标准规范:
            copy_file(folder_path,file,r'D:\Pan_xfusion_worksapce\1.《《任务》》\7月\1.知识库上传\7.w3\w3_知识库(强匹配)\ISC+通用-标准规范-内部公开')
        
        if find_V(file) in 业务操作指导书:
            copy_file(folder_path,file,r'D:\Pan_xfusion_worksapce\1.《《任务》》\7月\1.知识库上传\7.w3\w3_知识库(强匹配)\ISC+通用-业务操作指导书-内部公开')
        
        if find_V(file) in 管理手册:
            copy_file(folder_path,file,r'D:\Pan_xfusion_worksapce\1.《《任务》》\7月\1.知识库上传\7.w3\w3_知识库(强匹配)\ISC+通用-管理手册-内部公开')
        
        if find_V(file) in FAQ_message:
            copy_file(folder_path,file,r'D:\Pan_xfusion_worksapce\1.《《任务》》\7月\1.知识库上传\7.w3\w3_知识库(强匹配)\ISC+通用-FAQ库-内部公开')
            
    print("分类保存内容已完成！")

# 在O列写入'是'或'否'
# def check_list(row):
#     folder_path = r'D:\Pan_xfusion_worksapce\1.《《任务》》\7月\1.知识库上传\7.w3\w3文件_总'
#     return '是' if row['文档名称（当前）'] in folder_path else '否'  # 请确保列名与文件中一致

#判断是否有表格中是否有实体文件
def is_physical_file():
    folder_path = r'D:\Pan_xfusion_worksapce\1.《《任务》》\7月\1.知识库上传\7.w3\w3文件_总'
    w3_file_names = os.listdir(folder_path)
    list_01 = [find_V(filename) for filename in w3_file_names]
    # 加载Excel文件和特定Sheet
    file_path = 'test01.xlsx'
    sheet_name = '线上线下收集模板'
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook[sheet_name]
    for row in worksheet.iter_rows(min_row=2, values_only=False):  # 假设第一行是标题
        e_value = row[4].value  # E列索引为4（基于0的索引）
        c_value = find_V_row(row[2].value)  # C列索引为2
        o_cell = row[14]  # O列索引为14

        if e_value == 'W3':
            # 判断C列的值是否在list_01中
            if c_value in list_01:
                o_cell.value = '是'
            else:
                o_cell.value = '否'

    # 将结果保存到新的Excel文件
    workbook.save('updated_file.xlsx')

#将表格中没有的文件转移
def move_no_file():
    # 定义Excel文件路径和Sheet名称
    excel_file_path = 'test01.xlsx'
    sheet_name = '线上线下收集模板'
    # 定义源文件夹和目标文件夹路径
    source_folder = r'D:\Pan_xfusion_worksapce\1.《《任务》》\7月\1.知识库上传\7.w3\w3文件_总'
    target_folder = r'D:\Pan_xfusion_worksapce\1.《《任务》》\7月\1.知识库上传\7.w3\w3_知识库(强匹配)\w3_表格没有的文件_v5'
    # 如果目标文件夹不存在则创建
    if not os.path.exists(target_folder):
        os.makedirs(target_folder)
    # 加载Excel文件和特定Sheet
    workbook = openpyxl.load_workbook(excel_file_path)
    worksheet = workbook[sheet_name]
    # 获取Sheet中C列的数据，跳过标题行
    c_values = set(find_V_row(row[2].value) for row in worksheet.iter_rows(min_row=2) if row[2].value is not None)

    # 遍历源文件夹中的文件
    for filename in os.listdir(source_folder):
        # 文件绝对路径
        file_path = os.path.join(source_folder, filename)
        # 如果文件是普通文件而不是文件夹
        if os.path.isfile(file_path):
            # 去除文件后缀名
            file_name_no_extension = find_V(filename)
            # 如果文件名不存在于c_values中
            if file_name_no_extension not in c_values:
                # 复制文件到目标文件夹
                target_file_path = os.path.join(target_folder, filename)
                shutil.copy(file_path, target_file_path)
                print(f'Copied: {filename} to {target_folder}')


# get_all_list()
# file_classify()

# is_physical_file()
# move_no_file()
